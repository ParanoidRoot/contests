# @Time  : 2019/5/18 17:30
# @Author: Root
# @File  : Topsis.py


import openpyxl
import os
import numpy as np
import paranoid_root.contests.mathematical_modeling.group_a.Utility as uti
import random
import math


class College(object):
    """
    # 在这个类中完成对大学的构建
    """
    id = ""
    name = ""
    totalClicks = 0
    monthClicks = 0
    weekClicks = 0
    location = ""
    isDirected = -1
    is985 = -1
    is211 = -1
    schoolType = ""
    majors = {}
    furtherStudyRate = 0.0
    workRate = 0.0
    schoolScore = 0.0

    def __init__(self, _id, _name, _totalClks, _monthClks, _weekClks, _location, _isDirected, _is985,
                 _is211, _schoolType, _majors,
                 futherRate=0.0, workRate=0.0, _schoolScore=0.0):
        self.id = _id
        self.name = _name
        self.totalClicks = _totalClks
        self.monthClicks = _monthClks
        self.weekClicks = _weekClks
        self.location = _location
        self.isDirected = _isDirected
        self.is985 = _is985
        self.is211 = _is211
        self.schoolType = _schoolType
        self.majors = _majors
        self.furtherStudyRate = futherRate
        self.workRate = workRate
        self.schoolScore = _schoolScore

    @classmethod
    def buildByXlsRow(cls, row, workSheet):
        """
        通过row来构建一个College对象
        :param row:
        :param workSheet:
        :return:
        """
        id = workSheet.cell(row, 1).value
        name = workSheet.cell(row, 2).value
        totalClks = int(workSheet.cell(row, 3).value)
        monthClks = int(workSheet.cell(row, 4).value)
        weekClks = int(workSheet.cell(row, 5).value)
        location = workSheet.cell(row, 6).value
        isDrcted = int(workSheet.cell(row, 7).value)
        is985 = int(workSheet.cell(row, 8).value)
        is211 = int(workSheet.cell(row, 9).value)
        schoolType = workSheet.cell(row, 10).value

        majorStr = workSheet.cell(row, 11).value
        majors = majorStr.split("、")
        majors = set([major for major in majors if major != None and major != ""])

        furtherStudyRate = 0.0
        workRate = 0.0
        schoolScore = 0.0

        if workSheet.cell(row, 12).value:
            furtherStudyRate = float(workSheet.cell(row, 12).value)
            workRate = float(workSheet.cell(row, 13).value)
            schoolScore = float(workSheet.cell(row, 14).value)

        t = College(id, name, totalClks, monthClks, weekClks, location,
                    isDrcted, is985, is211,
                    schoolType, majors, furtherStudyRate, workRate,
                    schoolScore)
        return t

    def __str__(self):
        return (self.id, self.name, self.location,
                self.totalClicks, self.monthClicks, self.weekClicks,
                "直属" if self.isDirected else "非直属",
                "985" if self.is985 else "非985",
                "211" if self.is211 else "非211",
                self.schoolType,
                str(self.majors)
                ).__str__()

    def getMostSimilarMajorValue(self, majorStr):
        calculator = uti.Majors(self.majors)
        ans = calculator.getTopNSimilarity(majorStr)[0][0]
        return ans


class Colleges(object):
    """
    # 这个类完成构建所有的大学
    """
    xlsPath = ""  # 注意是absPath
    workBook = None
    workSheet = None
    collegeDict = None

    def __init__(self, xlsPath=os.path.join(os.getcwd(), r"data\allCollegeInfos.xlsx")):
        self.xlsPath = xlsPath
        self.workBook = openpyxl.load_workbook(self.xlsPath)
        self.workSheet = self.workBook.worksheets[0]
        self.collegeDict = dict()
        maxRow = self.workSheet.max_row
        for row in range(1, maxRow + 1):
            college = College.buildByXlsRow(row, self.workSheet)
            self.collegeDict[college.name] = self.collegeDict[college.id] = college

    def getCollegeByName(self, collegeName):
        return self.collegeDict[collegeName]

    def getCollegeById(self, collegeId):
        return self.collegeDict[collegeId]


class Student(object):
    rank = -1
    wantedMajors = {}

    def __init__(self, rank, wantedMajors):
        self.rank = rank
        self.wantedMajors = wantedMajors

    def __str__(self):
        return (self.rank, str(self.wantedMajors)).__str__()


class Topsis(object):
    collegeNumber = 0
    student = None
    allColleges = None
    possibleColleges = None
    rMatrix = None
    weightVector = None
    promptStr = "  专业匹配度 就业深造   整体实力   地理位置   社会认可度 专业热度   "

    def __init__(self, byMan, collegeNumber, student: Student, schools):
        self.collegeNumber = collegeNumber
        self.student = student
        self.allColleges = Colleges()
        self.possibleColleges = []
        self.pickColleges(self.pickByFixedColleges, schools)
        self.buildRMatrix()
        if not byMan :
            self.weightVector = self.buildWeightVector()
        else :
            print("请输入一个6 * 6的矩阵 : ")
            self.weightVector = np.zeros((6, 6) )
            lines = [ input() for i in range(6)]
            for row, line in enumerate(lines) :
                parts = line.strip().split()
                for col, part in enumerate(parts) :
                    value = float(eval(part))
                    self.weightVector[row][col] = value
            self.weightVector = self.getWeightVector(self.weightVector, 6)



    def buildRMatrix(self):
        """
        通过学生的信息构建出一个R矩阵
        :return:
        """
        self.rMatrix = np.zeros((self.collegeNumber, 6) )
        print(self.rMatrix)
        self.initializeColZero()
        self.initializeColOne()
        self.initializeColTwo()
        self.initializeColThree()
        self.initializeColFour()
        self.initializeColFive()

        print("after col normalized : ")
        self.rMatrix = self.normalizeMatrixByCols(self.rMatrix)
        print(self.rMatrix)
        print("total normalized : ")
        print(self.promptStr)
        self.rMatrix = self.normalizeMatrix(self.rMatrix)
        print(self.rMatrix)
        for i, college in enumerate(self.possibleColleges) :
            print(i, college.name)

    def buildWeightVector(self):
        """
        生成v矩阵并返回
        :return:
        """
        input("\r\n请学生对各个因素之比进行模糊打分 : ")
        print("学生打分为 :")
        print(self.promptStr)
        t = self.buildWeightMatrix(6)
        print(t)
        t = self.getWeightVector(t, 6)
        print("生成的权重向量为 : ")
        print(t)
        return t

    def pickColleges(self, picker, *args):
        """
        传入一个pciker函数
        :param picker:
        :return:
        """
        names = picker(*args)
        for name in names:
            self.possibleColleges.append(self.allColleges.collegeDict[name])

    def pickByFixedColleges(self, colleges):
        ans = set(colleges)
        if len(ans) != self.collegeNumber:
            raise Exception("学校列表出现大小错误")
        return ans

    def initializeColZero(self):
        """
        初始化第一列
        :return:
        """
        if not self.student.wantedMajors :
            uti.Majors.initializeClass()
            for row, college in enumerate(self.possibleColleges) :
                self.rMatrix[row][0] = 1.0
        else :
            for row, college in enumerate(self.possibleColleges):
                self.rMatrix[row][0] = max([college.getMostSimilarMajorValue(loved) for loved in self.student.wantedMajors])

    def initializeColOne(self):
        """
        保研率与深造率
        :return:
        """
        choice = input("study(1) or job(0) : ")
        if choice == "1":
            for row, college in enumerate(self.possibleColleges):
                self.rMatrix[row][1] = college.furtherStudyRate
        else:
            for row, college in enumerate(self.possibleColleges):
                self.rMatrix[row][1] = college.workRate

    def initializeColTwo(self):
        """
        :return:
        """
        for row, college in enumerate(self.possibleColleges):
            self.rMatrix[row][2] = college.schoolScore

    def initializeColThree(self):
        """

        :return:
        """
        print("请对以下地理位置进行排序 : ".center(25, "*"))
        locations = list(set([college.location for college in self.possibleColleges]))
        print("请选择 : ")
        print(locations)
        input()
        decisions = self.buildWeightMatrix(len(locations), 9)
        print("选择为 :\r\n", decisions)
        weights = self.getWeightVector(decisions, len(locations))
        print("权重向量为 :\r\n", weights)
        for row, location in enumerate(locations):
            for i, college in enumerate(self.possibleColleges):
                if college.location == location:
                    self.rMatrix[i][3] = weights[row]

    def initializeColFour(self):
        """
        初始化第四行，clicks
        :return:
        """
        for row, college in enumerate(self.possibleColleges):
            self.rMatrix[row][4] = college.totalClicks

    def initializeColFive(self):
        """
        学校的特色专业与热点行业之间的契合度
        :return:
        """

        for row, college in enumerate(self.possibleColleges):
            self.rMatrix[row][5] = max(
                [college.getMostSimilarMajorValue(hotMajor) for hotMajor in uti.Majors.hotMajors])

    def buildWeightMatrix(self, n: int, bound=9):
        """
        生成 n*n 的正互反矩阵
        :param n:
        :return:
        """
        mat = np.zeros((n, n))
        for row in range(n):
            for col in range(row, n):
                if row == col:
                    mat[row][col] = 1.0
                else:
                    t = random.randint(0, 1)
                    if t:
                        mat[row][col] = 1.0 * random.randint(1, bound)
                        mat[col][row] = 1 / mat[row][col]
                    else:
                        mat[col][row] = 1.0 * random.randint(1, bound)
                        mat[row][col] = 1 / mat[col][row]
        return mat

    def getWeightVector(self, mat: np.array, n):
        """
        获取 n*n 的正互反矩阵的权重向量
        :param mat:
        :param n:
        :return:
        """
        vector = np.sum(mat, axis=1).reshape((-1, 1))
        vector = vector / n
        vector = self.normalizeMatrix(vector)
        return vector

    def normalizeMatrix(self, matrix):
        length = math.sqrt(np.sum(matrix ** 2))
        return matrix / length

    def calculate(self) :
        v = np.zeros((6, 6) )
        for i in range(6) :
            v[i][i] = self.weightVector[i]
        temp = self.rMatrix.dot(v)
        minVector = np.min(temp, axis=0)
        maxVector = np.max(temp, axis=0)
        print("temp is \r\n", temp)
        print("minVector is \r\n", minVector)
        print("maxVector is \r\n", maxVector)
        ansList = []
        for row in range(self.collegeNumber) :
            vector = temp[row, :]
            D2Min, D2Max = self.getD1D2(minVector, maxVector, vector)
            k = D2Min / (D2Max + D2Min)
            ansList.append( (k, self.possibleColleges[row] ) )
        return ansList

    def getFinal(self, n : int) :
        ansList = self.calculate()
        t = sorted(ansList, key= lambda x : x[0],reverse=True)
        print("\r\n", "".center(25, "*") )
        for e in t :
            print(str(e[0]), str(e[1]) )
        print("\r\n推荐的决策为 : ")
        for i in range(n) :
            print(str(i + 1).center(25, "*"))
            print(str(t[i][0]), str(t[i][1]))

    def getAns2(self) :
        for col in range(6) :
            tempList = []
            tempCol = self.rMatrix[ : , col]
            for row in range(self.collegeNumber) :
                tempList.append( (tempCol[row], row) )
            tempList = sorted(tempList, key= lambda x: x[0], reverse=True)
            print()
            if col == 0 :
                print("理想专业与学校特色专业的匹配度 : ".center(25, "*"))
            elif col == 1 :
                print("就业深造机会 : ".center(25, "*"))
            elif col == 2 :
                print("整体实力 : ".center(25, "*"))
            elif col == 3 :
                print("地理位置 : ".center(25, "*"))
            elif col == 4 :
                print("社会认可度 : ".center(25, "*"))
            elif col == 5 :
                print("学校专业与热门专业的契合度 : ".center(25, "*"))
            for i in range(10) :
                print(tempList[i][0], self.possibleColleges[tempList[i][1]].name)

    def getD1D2(self, minVector, maxVector, temp) :
        return (self.calculateDistance(minVector, temp), self.calculateDistance(maxVector, temp) )

    def calculateDistance(self, vector1, vector2) :
        delta = vector1 - vector2
        delta = delta ** 2
        sumDelta = np.sum(delta)
        return math.sqrt(sumDelta)

    def normalizeMatrixByCols(self, matrix : np.array) :
        """
        对矩阵的每个列进行归一化
        :param matrix:
        :return:
        """
        for col in range(6) :
            colVector = matrix[: , col ]
            colVector = self.normalizeMatrix(colVector)
            for raw in range(self.collegeNumber) :
                matrix[raw][col] = colVector[raw]
        return matrix








